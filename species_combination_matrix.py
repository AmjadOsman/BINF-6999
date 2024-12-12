import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the CSV file
file_path = r''
df = pd.read_csv(file_path)

# Extract the relevant columns for species and the combination
species_col = 'rank_species'  # Column for species
media_col = 'Media'
condition_col = 'Condition'
chemostat_col = 'Chemostat'
atmosphere_col = 'Atmosphere'

# Drop rows where any of the required columns are missing
df = df.dropna(subset=[species_col, media_col, condition_col, chemostat_col, atmosphere_col])

# Create a new column that combines media, condition, chemostat, and atmosphere
df['Combination'] = df[media_col] + '_' + df[condition_col] + '_' + df[chemostat_col] + '_' + df[atmosphere_col]

# Get unique species and combinations
unique_species = df[species_col].unique()
unique_combinations = df['Combination'].unique()

# Create a DataFrame for the matrix
matrix_df = pd.DataFrame(index=unique_species, columns=unique_combinations)

# Fill the matrix with 'Y' or 'N'
for spec in unique_species:
    for combo in unique_combinations:
        if ((df[species_col] == spec) & (df['Combination'] == combo)).any():
            matrix_df.at[spec, combo] = 'Y'
        else:
            matrix_df.at[spec, combo] = 'N'

# Save the matrix to an Excel file using openpyxl to apply conditional formatting
output_file_path = r''
with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    matrix_df.to_excel(writer, sheet_name='Species Combination Matrix')
    
    # Get the workbook and the sheet to apply conditional formatting
    workbook = writer.book
    sheet = workbook['Species Combination Matrix']
    
    # Define cell colors
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green for 'Y'
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')    # Red for 'N'
    
    # Apply conditional formatting
    for row in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            if cell.value == 'Y':
                cell.fill = green_fill
            elif cell.value == 'N':
                cell.fill = red_fill

print(f"Species-combination matrix with color coding saved to {output_file_path}")
